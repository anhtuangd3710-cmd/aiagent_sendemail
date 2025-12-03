"""
Chatbot Service - AI-powered chatbot for email analytics
Allows users to ask questions about their email data
Generates charts, statistics, and Excel exports
"""
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import io
import base64

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ChatbotService:
    """Service for AI-powered chatbot with email analytics"""
    
    def __init__(self, database, ai_agent):
        self.database = database
        self.ai_agent = ai_agent
    
    def _get_all_emails_admin(self) -> List[Dict]:
        """Get all emails from all users (admin only)"""
        try:
            emails = self.database.query_raw(
                """SELECT e.*, u.username as user_name, u.email as user_email 
                   FROM emails e 
                   LEFT JOIN users u ON e.user_id = u.id 
                   ORDER BY e.sent_at DESC"""
            )
            return emails or []
        except Exception as e:
            logger.error(f"Error getting all emails for admin: {e}")
            return []
    
    def _get_all_cv_evaluations_admin(self) -> List[Dict]:
        """Get all CV evaluations from all users (admin only)"""
        try:
            evaluations = self.database.query_raw(
                """SELECT cv.*, u.username as user_name, u.email as user_email 
                   FROM cv_evaluations cv 
                   LEFT JOIN users u ON cv.user_id = u.id 
                   ORDER BY cv.created_at DESC"""
            )
            return evaluations or []
        except Exception as e:
            logger.error(f"Error getting all CV evaluations for admin: {e}")
            return []
    
    def _get_user_statistics_admin(self) -> Dict:
        """Get user statistics (admin only)"""
        try:
            users = self.database.query_raw(
                """SELECT id, username, email, full_name, role, is_active, 
                          created_at, last_login 
                   FROM users ORDER BY created_at DESC"""
            )
            
            total_users = len(users) if users else 0
            active_users = sum(1 for u in users if u.get('is_active')) if users else 0
            admin_count = sum(1 for u in users if u.get('role') == 'admin') if users else 0
            
            return {
                'total_users': total_users,
                'active_users': active_users,
                'admin_count': admin_count,
                'user_count': total_users - admin_count,
                'users_list': users[:10] if users else []  # Top 10 recent users
            }
        except Exception as e:
            logger.error(f"Error getting user statistics: {e}")
            return {}
        
    def get_email_statistics(self, user_id: int, time_range: str = None, is_admin: bool = False) -> Dict:
        """Get comprehensive email statistics for a user (or all users if admin)"""
        try:
            # Get emails - all for admin, user-specific otherwise
            if is_admin:
                emails = self._get_all_emails_admin()
            else:
                emails = self.database.get_all_emails(user_id)
            
            # Filter by time range if specified
            if time_range:
                emails = self._filter_by_time_range(emails, time_range)
            
            # Calculate statistics
            total_sent = len(emails)
            responded = sum(1 for e in emails if e.get('response_received'))
            pending = total_sent - responded
            
            # Calculate response rate
            response_rate = (responded / total_sent * 100) if total_sent > 0 else 0
            
            # Analyze sentiments from responses
            sentiments = {'positive': 0, 'negative': 0, 'neutral': 0}
            decisions = {'agreed': 0, 'disagreed': 0, 'undecided': 0, 'needs_more_info': 0}
            
            for email in emails:
                if email.get('response_received') and email.get('analysis'):
                    analysis = email['analysis']
                    if isinstance(analysis, str):
                        try:
                            analysis = json.loads(analysis)
                        except:
                            continue
                    
                    sentiment = analysis.get('sentiment', '').lower()
                    if 'positive' in sentiment or 'tích cực' in sentiment:
                        sentiments['positive'] += 1
                    elif 'negative' in sentiment or 'tiêu cực' in sentiment:
                        sentiments['negative'] += 1
                    else:
                        sentiments['neutral'] += 1
                    
                    decision = analysis.get('decision', '').lower()
                    if 'agree' in decision or 'đồng ý' in decision:
                        decisions['agreed'] += 1
                    elif 'disagree' in decision or 'không đồng ý' in decision or 'từ chối' in decision:
                        decisions['disagreed'] += 1
                    elif 'more_info' in decision or 'thêm thông tin' in decision:
                        decisions['needs_more_info'] += 1
                    else:
                        decisions['undecided'] += 1
            
            # Get emails by date for time series
            emails_by_date = self._group_emails_by_date(emails)
            
            # Get top recipients
            top_recipients = self._get_top_recipients(emails)
            
            # Get emails by purpose
            emails_by_purpose = self._group_by_purpose(emails)
            
            return {
                'total_sent': total_sent,
                'responded': responded,
                'pending': pending,
                'failed': 0,  # Could track failed emails if needed
                'response_rate': round(response_rate, 2),
                'sentiments': sentiments,
                'decisions': decisions,
                'emails_by_date': emails_by_date,
                'top_recipients': top_recipients,
                'emails_by_purpose': emails_by_purpose,
                'time_range': time_range or 'all',
                'is_admin_view': is_admin
            }
        except Exception as e:
            logger.error(f"Error getting email statistics: {e}")
            return {}
    
    def get_cv_statistics(self, user_id: int, is_admin: bool = False) -> Dict:
        """Get CV evaluation statistics (all users if admin)"""
        try:
            if is_admin:
                evaluations = self._get_all_cv_evaluations_admin()
            else:
                evaluations = self.database.get_all_cv_evaluations(user_id)
            
            total = len(evaluations)
            qualified = sum(1 for e in evaluations if e.get('is_qualified'))
            not_qualified = total - qualified
            emails_sent = sum(1 for e in evaluations if e.get('email_sent'))
            
            # Score distribution
            score_ranges = {
                '0-50': 0,
                '51-70': 0,
                '71-84': 0,
                '85-100': 0
            }
            
            for ev in evaluations:
                score = ev.get('overall_score', 0)
                if score <= 50:
                    score_ranges['0-50'] += 1
                elif score <= 70:
                    score_ranges['51-70'] += 1
                elif score <= 84:
                    score_ranges['71-84'] += 1
                else:
                    score_ranges['85-100'] += 1
            
            # Average score
            avg_score = sum(e.get('overall_score', 0) for e in evaluations) / total if total > 0 else 0
            
            # By job title
            by_job_title = {}
            for ev in evaluations:
                job = ev.get('job_title', 'Không xác định')
                if job not in by_job_title:
                    by_job_title[job] = {'total': 0, 'qualified': 0}
                by_job_title[job]['total'] += 1
                if ev.get('is_qualified'):
                    by_job_title[job]['qualified'] += 1
            
            return {
                'total': total,
                'qualified': qualified,
                'not_qualified': not_qualified,
                'qualification_rate': round(qualified / total * 100, 2) if total > 0 else 0,
                'emails_sent': emails_sent,
                'average_score': round(avg_score, 2),
                'score_distribution': score_ranges,
                'by_job_title': by_job_title,
                'is_admin_view': is_admin
            }
        except Exception as e:
            logger.error(f"Error getting CV statistics: {e}")
            return {}
    
    def _filter_by_time_range(self, emails: List[Dict], time_range: str) -> List[Dict]:
        """Filter emails by time range"""
        now = datetime.now()
        
        ranges = {
            'today': timedelta(days=1),
            'week': timedelta(weeks=1),
            'month': timedelta(days=30),
            '3months': timedelta(days=90),
            'year': timedelta(days=365)
        }
        
        if time_range not in ranges:
            return emails
        
        cutoff = now - ranges[time_range]
        
        filtered = []
        for email in emails:
            try:
                sent_at = email.get('sent_at')
                if sent_at:
                    if isinstance(sent_at, str):
                        # Try multiple formats
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d']:
                            try:
                                sent_date = datetime.strptime(sent_at[:19], fmt)
                                break
                            except:
                                continue
                        else:
                            continue
                    else:
                        sent_date = sent_at
                    
                    if sent_date >= cutoff:
                        filtered.append(email)
            except Exception as e:
                logger.warning(f"Error parsing date: {e}")
                continue
        
        return filtered
    
    def _group_emails_by_date(self, emails: List[Dict]) -> List[Dict]:
        """Group emails by date for time series chart"""
        by_date = {}
        
        for email in emails:
            try:
                sent_at = email.get('sent_at', '')
                if sent_at:
                    date_str = str(sent_at)[:10]
                    if date_str not in by_date:
                        by_date[date_str] = {'sent': 0, 'responded': 0}
                    by_date[date_str]['sent'] += 1
                    if email.get('response_received'):
                        by_date[date_str]['responded'] += 1
            except:
                continue
        
        # Sort by date and return as list
        sorted_dates = sorted(by_date.items())
        return [{'date': d, 'sent': v['sent'], 'responded': v['responded']} for d, v in sorted_dates]
    
    def _get_top_recipients(self, emails: List[Dict], limit: int = 5) -> List[Dict]:
        """Get top email recipients"""
        recipients = {}
        
        for email in emails:
            recipient = email.get('recipient_email', '')
            name = email.get('recipient_name', '')
            if recipient:
                if recipient not in recipients:
                    recipients[recipient] = {'email': recipient, 'name': name, 'count': 0, 'responded': 0}
                recipients[recipient]['count'] += 1
                if email.get('response_received'):
                    recipients[recipient]['responded'] += 1
        
        # Sort by count and return top N
        sorted_recipients = sorted(recipients.values(), key=lambda x: x['count'], reverse=True)
        return sorted_recipients[:limit]
    
    def _group_by_purpose(self, emails: List[Dict]) -> Dict:
        """Group emails by purpose/category"""
        purposes = {}
        
        for email in emails:
            purpose = email.get('purpose', 'Khác')
            # Truncate long purposes
            if len(purpose) > 50:
                purpose = purpose[:47] + '...'
            
            if purpose not in purposes:
                purposes[purpose] = {'count': 0, 'responded': 0}
            purposes[purpose]['count'] += 1
            if email.get('response_received'):
                purposes[purpose]['responded'] += 1
        
        return purposes
    
    def _get_guide_content(self) -> str:
        """Get comprehensive guide content for the chatbot"""
        return """
=== HƯỚNG DẪN SỬ DỤNG HỆ THỐNG EMAIL AI AGENT ===

📧 **1. SOẠN VÀ GỬI EMAIL**
- Vào menu "Soạn Email" từ thanh bên trái
- Nhập thông tin người gửi (tên, chức vụ, công ty)
- Nhập thông tin người nhận (email, tên)
- Mô tả mục đích email, AI sẽ tự động soạn nội dung
- Xem trước và gửi email

📋 **2. ĐÁNH GIÁ CV ỨNG VIÊN**
- Vào menu "Đánh giá CV"
- Upload file CV (PDF, DOCX, TXT) hoặc paste nội dung
- Nhập tiêu chí tuyển dụng và vị trí
- AI sẽ đánh giá và cho điểm từ 0-100
- Nếu điểm ≥85%, tự động gửi thư mời phỏng vấn

📊 **3. GIÁM SÁT PHẢN HỒI**
- Vào menu "Giám sát phản hồi"
- Hệ thống tự động kiểm tra email phản hồi
- AI phân tích cảm xúc và quyết định từ phản hồi
- Thông báo realtime qua WebSocket

📈 **4. PHÂN TÍCH YOUTUBE**
- Vào menu "Phân tích YouTube"
- Nhập URL kênh YouTube
- Xem thống kê: subscribers, views, videos
- AI ước tính thu nhập và phân tích xu hướng

🤖 **5. CHATBOT THỐNG KÊ** (Bạn đang ở đây!)
- Hỏi về thống kê email, CV
- Yêu cầu vẽ biểu đồ
- Xuất dữ liệu ra Excel
- Hỏi hướng dẫn sử dụng

=== HƯỚNG DẪN THIẾT LẬP API VÀ EMAIL ===

🔑 **THIẾT LẬP EMAIL GỬI (App Password)**

📧 **Gmail:**
1. Truy cập https://myaccount.google.com/security
2. Bật "Xác minh 2 bước" (2-Step Verification)
3. Truy cập https://myaccount.google.com/apppasswords
4. Chọn "Select app" → "Mail"
5. Chọn "Select device" → "Other" → Nhập "Email AI Agent"
6. Nhấn "Generate" → Sao chép mật khẩu 16 ký tự
7. Vào Hồ sơ & API Keys → Nhập email và App Password

📧 **Outlook/Hotmail:**
1. Truy cập https://account.microsoft.com/security
2. Bật Xác minh 2 bước
3. Vào "App passwords" → Tạo mật khẩu mới
4. Sao chép và sử dụng mật khẩu được tạo
5. SMTP: smtp.office365.com, Port: 587
6. IMAP: outlook.office365.com, Port: 993

📧 **Yahoo Mail:**
1. Truy cập https://login.yahoo.com/account/security
2. Bật Xác minh 2 bước
3. Nhấn "Generate app password"
4. Chọn "Other app", nhập tên, nhấn "Generate"
5. SMTP: smtp.mail.yahoo.com, Port: 587
6. IMAP: imap.mail.yahoo.com, Port: 993

🤖 **THIẾT LẬP API AI**

🔹 **Google Gemini API (Miễn phí):**
1. Truy cập https://makersuite.google.com/app/apikey
2. Đăng nhập bằng tài khoản Google
3. Nhấn "Create API Key"
4. Sao chép API Key
5. Vào Hồ sơ & API Keys → Chọn Gemini → Dán API Key

🔹 **Azure OpenAI:**
1. Truy cập https://portal.azure.com
2. Tạo Azure OpenAI resource
3. Vào Keys and Endpoint → Sao chép KEY 1 và Endpoint
4. Deploy model GPT-4 hoặc GPT-3.5
5. Vào Hồ sơ & API Keys → Chọn Azure → Nhập thông tin

🔹 **OpenAI API:**
1. Truy cập https://platform.openai.com/api-keys
2. Nhấn "Create new secret key"
3. Sao chép API Key (chỉ hiện 1 lần!)
4. Vào Hồ sơ & API Keys → Chọn OpenAI → Dán API Key

=== CẤU HÌNH SMTP/IMAP PHỔ BIẾN ===

| Provider | SMTP Host | SMTP Port | IMAP Host | IMAP Port |
|----------|-----------|-----------|-----------|-----------|
| Gmail | smtp.gmail.com | 587 | imap.gmail.com | 993 |
| Outlook | smtp.office365.com | 587 | outlook.office365.com | 993 |
| Yahoo | smtp.mail.yahoo.com | 587 | imap.mail.yahoo.com | 993 |
| iCloud | smtp.mail.me.com | 587 | imap.mail.me.com | 993 |

=== MẸO SỬ DỤNG ===

💡 **Tips:**
- Luôn test gửi email cho chính mình trước
- Kiểm tra thư mục Spam nếu không nhận được email
- Sử dụng mô tả chi tiết để AI soạn email tốt hơn
- Cập nhật tiêu chí tuyển dụng rõ ràng khi đánh giá CV
- Xuất Excel để lưu trữ và báo cáo

⚠️ **Lưu ý bảo mật:**
- Không chia sẻ App Password với người khác
- API Key nên được bảo mật
- Đổi mật khẩu định kỳ
"""

    def _is_guide_query(self, query: str) -> bool:
        """Check if query is asking for help/guide"""
        guide_keywords = [
            'hướng dẫn', 'cách sử dụng', 'cách dùng', 'làm sao', 'làm thế nào',
            'thiết lập', 'cài đặt', 'setup', 'config', 'cấu hình',
            'api', 'api key', 'gemini', 'openai', 'azure',
            'app password', 'mật khẩu ứng dụng', 'password',
            'gmail', 'outlook', 'yahoo', 'email gửi',
            'smtp', 'imap', 'port',
            'help', 'trợ giúp', 'hỗ trợ',
            'bắt đầu', 'getting started', 'tutorial',
            'tính năng', 'chức năng', 'feature'
        ]
        query_lower = query.lower()
        return any(keyword in query_lower for keyword in guide_keywords)

    def process_query(self, user_id: int, query: str, user_settings: Dict = None, 
                       is_admin: bool = False, session_id: int = None) -> Dict:
        """Process user query and return appropriate response with data"""
        try:
            # Get or create session
            if session_id is None:
                session_id = self.database.get_or_create_active_session(user_id)
            
            # Save user message
            self.database.save_chat_message(session_id, user_id, 'user', query)
            
            # Check if this is a guide/help query
            is_guide_query = self._is_guide_query(query)
            
            # Get user data - admin can see all data
            email_stats = self.get_email_statistics(user_id, is_admin=is_admin)
            cv_stats = self.get_cv_statistics(user_id, is_admin=is_admin)
            user_stats = self._get_user_statistics_admin() if is_admin else None
            
            # Build context for AI
            admin_context = ""
            if is_admin and user_stats:
                admin_context = f"""
=== THỐNG KÊ NGƯỜI DÙNG (CHỈ ADMIN) ===
- Tổng số người dùng: {user_stats.get('total_users', 0)}
- Người dùng đang hoạt động: {user_stats.get('active_users', 0)}
- Số admin: {user_stats.get('admin_count', 0)}
- Số user thường: {user_stats.get('user_count', 0)}

⚠️ LƯU Ý: Bạn đang xem dữ liệu TOÀN BỘ HỆ THỐNG với quyền Admin.
"""
            
            # Add guide content if user is asking for help
            guide_context = ""
            if is_guide_query:
                guide_context = self._get_guide_content()
            
            view_note = "(Dữ liệu toàn hệ thống - Admin View)" if is_admin else "(Dữ liệu cá nhân)"
            
            context = f"""
Bạn là chatbot trợ lý thông minh của hệ thống Email AI Agent. Bạn có thể:
1. Trả lời câu hỏi về thống kê email và CV
2. Hướng dẫn sử dụng hệ thống
3. Hướng dẫn thiết lập API (Gemini, OpenAI, Azure)
4. Hướng dẫn cấu hình email (Gmail, Outlook, Yahoo App Password)
5. Vẽ biểu đồ và xuất Excel

{guide_context}
{admin_context}

=== THỐNG KÊ EMAIL {view_note} ===
- Tổng số email đã gửi: {email_stats.get('total_sent', 0)}
- Số email đã nhận phản hồi: {email_stats.get('responded', 0)}
- Số email đang chờ phản hồi: {email_stats.get('pending', 0)}
- Tỷ lệ phản hồi: {email_stats.get('response_rate', 0)}%

Phân tích cảm xúc phản hồi:
- Tích cực: {email_stats.get('sentiments', {}).get('positive', 0)}
- Tiêu cực: {email_stats.get('sentiments', {}).get('negative', 0)}
- Trung tính: {email_stats.get('sentiments', {}).get('neutral', 0)}

Quyết định từ phản hồi:
- Đồng ý: {email_stats.get('decisions', {}).get('agreed', 0)}
- Từ chối: {email_stats.get('decisions', {}).get('disagreed', 0)}
- Chưa quyết định: {email_stats.get('decisions', {}).get('undecided', 0)}
- Cần thêm thông tin: {email_stats.get('decisions', {}).get('needs_more_info', 0)}

Top 5 người nhận email:
{json.dumps(email_stats.get('top_recipients', []), ensure_ascii=False, indent=2)}

=== THỐNG KÊ ĐÁNH GIÁ CV {view_note} ===
- Tổng số CV đã đánh giá: {cv_stats.get('total', 0)}
- Số ứng viên đạt yêu cầu (≥85%): {cv_stats.get('qualified', 0)}
- Số ứng viên chưa đạt: {cv_stats.get('not_qualified', 0)}
- Tỷ lệ đạt: {cv_stats.get('qualification_rate', 0)}%
- Điểm trung bình: {cv_stats.get('average_score', 0)}
- Số email mời phỏng vấn đã gửi: {cv_stats.get('emails_sent', 0)}

Phân bố điểm CV:
{json.dumps(cv_stats.get('score_distribution', {}), ensure_ascii=False, indent=2)}

=== CÂU HỎI CỦA NGƯỜI DÙNG ===
{query}

=== HƯỚNG DẪN TRẢ LỜI ===
- Trả lời tự nhiên, thân thiện bằng tiếng Việt
- Nếu hỏi về hướng dẫn/thiết lập: Cung cấp hướng dẫn chi tiết từng bước
- Nếu hỏi về thống kê: Cung cấp số liệu và phân tích
- Nếu hỏi về biểu đồ/xuất file: Đề cập bạn có thể hỗ trợ
- Sử dụng emoji để làm nổi bật thông tin quan trọng
- Định dạng với bullet points cho dễ đọc
"""
            
            # Determine if user wants a chart or export
            query_lower = query.lower()
            wants_chart = any(word in query_lower for word in ['biểu đồ', 'chart', 'vẽ', 'đồ thị', 'graph', 'visual', 'thống kê'])
            wants_export = any(word in query_lower for word in ['excel', 'xuất', 'export', 'download', 'tải'])
            
            # Determine chart type and get chart data if needed
            chart_type = self._determine_chart_type(query_lower)
            chart_data = None
            if wants_chart:
                chart_data = self.get_chart_data(user_id, chart_type, is_admin=is_admin)
            
            # Generate AI response
            ai_response = self._generate_ai_response(context, user_settings)
            
            response = {
                'success': True,
                'message': ai_response,
                'data': {
                    'email_stats': email_stats,
                    'cv_stats': cv_stats,
                    'user_stats': user_stats if is_admin else None
                },
                'is_admin': is_admin,
                'show_chart': wants_chart,
                'show_export': wants_export,
                'chart_type': chart_type,
                'chart_data': chart_data,
                'session_id': session_id,
                'timestamp': datetime.now().isoformat()
            }
            
            # Save bot response
            self.database.save_chat_message(
                session_id, user_id, 'bot', ai_response,
                has_chart=wants_chart, chart_type=chart_type,
                chart_data=chart_data, has_export=wants_export
            )
            
            # Update session title if this is the first message
            self.database.generate_session_title(session_id, user_id)
            
            return response
            
        except Exception as e:
            logger.error(f"Error processing query: {e}")
            return {
                'success': False,
                'message': f'Xin lỗi, đã có lỗi xảy ra: {str(e)}',
                'data': {}
            }
    
    def _generate_ai_response(self, context: str, user_settings: Dict = None) -> str:
        """Generate AI response for the query"""
        try:
            import google.generativeai as genai
            from config.settings import GEMINI_API_KEY, GEMINI_MODEL
            
            # Use user's API key if available
            api_key = GEMINI_API_KEY
            if user_settings and user_settings.get('gemini_api_key'):
                api_key = user_settings['gemini_api_key']
            
            if not api_key:
                return "Xin lỗi, chưa cấu hình API key cho AI. Vui lòng kiểm tra cài đặt."
            
            genai.configure(api_key=api_key)
            
            model = genai.GenerativeModel(
                model_name=GEMINI_MODEL,
                generation_config={
                    "temperature": 0.7,
                    "top_p": 0.95,
                    "max_output_tokens": 2048,
                }
            )
            
            response = model.generate_content(context)
            return response.text
            
        except Exception as e:
            logger.error(f"Error generating AI response: {e}")
            # Fallback to basic response
            return self._generate_basic_response(context)
    
    def _generate_basic_response(self, context: str) -> str:
        """Generate basic response without AI"""
        # Extract stats from context
        return """Dựa trên dữ liệu của bạn:
        
📊 **Tổng quan Email:**
- Xem phần thống kê bên phải để biết chi tiết
- Bạn có thể xem biểu đồ hoặc xuất file Excel

💡 **Gợi ý:**
- Hỏi "Vẽ biểu đồ số email theo ngày" để xem biểu đồ
- Hỏi "Xuất Excel" để tải dữ liệu

Tôi có thể giúp gì thêm cho bạn?"""
    
    def _determine_chart_type(self, query: str) -> str:
        """Determine the appropriate chart type based on query"""
        query_lower = query.lower()
        
        # Pie chart keywords
        if any(word in query_lower for word in ['pie', 'tròn', 'cảm xúc', 'sentiment', 'phần trăm']):
            return 'pie'
        
        # Bar chart keywords
        elif any(word in query_lower for word in ['bar', 'cột', 'người nhận', 'recipient', 'top']):
            return 'bar'
        
        # Line chart keywords
        elif any(word in query_lower for word in ['line', 'đường', 'xu hướng', 'trend', 'theo ngày', 'thời gian']):
            return 'line'
        
        # Doughnut chart keywords
        elif any(word in query_lower for word in ['doughnut', 'donut', 'trạng thái', 'status']):
            return 'doughnut'
        
        # CV related charts
        elif any(word in query_lower for word in ['cv', 'ứng viên', 'điểm cv', 'phân bố điểm']):
            return 'cv'
        
        # CV qualification
        elif any(word in query_lower for word in ['đạt yêu cầu', 'qualified', 'tỷ lệ đạt']):
            return 'cv_qualified'
        
        # Decision analysis
        elif any(word in query_lower for word in ['quyết định', 'decision', 'đồng ý', 'từ chối']):
            return 'decision'
        
        # Radar/performance chart
        elif any(word in query_lower for word in ['radar', 'hiệu suất', 'performance', 'tổng quan hiệu suất']):
            return 'radar'
        
        else:
            return 'overview'  # Default to doughnut overview
    
    def generate_excel_data(self, user_id: int, data_type: str = 'all', is_admin: bool = False) -> bytes:
        """Generate Excel file with email/CV data (all data if admin)"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            
            wb = openpyxl.Workbook()
            
            # Styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if data_type in ['all', 'emails']:
                # Email sheet
                ws_emails = wb.active
                ws_emails.title = 'Emails đã gửi' + (' (Tất cả)' if is_admin else '')
                
                if is_admin:
                    emails = self._get_all_emails_admin()
                else:
                    emails = self.database.get_all_emails(user_id)
                
                # Headers - add user column for admin
                if is_admin:
                    headers = ['ID', 'User', 'Ngày gửi', 'Người nhận', 'Email', 'Tiêu đề', 'Mục đích', 'Đã phản hồi', 'Cảm xúc', 'Quyết định']
                else:
                    headers = ['ID', 'Ngày gửi', 'Người nhận', 'Email', 'Tiêu đề', 'Mục đích', 'Đã phản hồi', 'Cảm xúc', 'Quyết định']
                for col, header in enumerate(headers, 1):
                    cell = ws_emails.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Data
                for row_num, email in enumerate(emails, 2):
                    analysis = email.get('analysis', {})
                    if isinstance(analysis, str):
                        try:
                            analysis = json.loads(analysis)
                        except:
                            analysis = {}
                    
                    col_offset = 1 if is_admin else 0
                    ws_emails.cell(row=row_num, column=1, value=email.get('id'))
                    if is_admin:
                        ws_emails.cell(row=row_num, column=2, value=email.get('user_name', email.get('user_email', '')))
                    ws_emails.cell(row=row_num, column=2 + col_offset, value=str(email.get('sent_at', ''))[:19])
                    ws_emails.cell(row=row_num, column=3 + col_offset, value=email.get('recipient_name', ''))
                    ws_emails.cell(row=row_num, column=4 + col_offset, value=email.get('recipient_email', ''))
                    ws_emails.cell(row=row_num, column=5 + col_offset, value=email.get('subject', ''))
                    ws_emails.cell(row=row_num, column=6 + col_offset, value=email.get('purpose', ''))
                    ws_emails.cell(row=row_num, column=7 + col_offset, value='Có' if email.get('response_received') else 'Không')
                    ws_emails.cell(row=row_num, column=8 + col_offset, value=analysis.get('sentiment', ''))
                    ws_emails.cell(row=row_num, column=9 + col_offset, value=analysis.get('decision', ''))
                
                # Auto-fit columns
                for col in ws_emails.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = min(len(str(cell.value)), 50)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_emails.column_dimensions[column].width = adjusted_width
            
            if data_type in ['all', 'cv']:
                # CV sheet
                ws_cv = wb.create_sheet('Đánh giá CV' + (' (Tất cả)' if is_admin else '')) if data_type == 'all' else wb.active
                if data_type != 'all':
                    ws_cv.title = 'Đánh giá CV' + (' (Tất cả)' if is_admin else '')
                
                if is_admin:
                    evaluations = self._get_all_cv_evaluations_admin()
                else:
                    evaluations = self.database.get_all_cv_evaluations(user_id)
                
                # Headers - add user column for admin
                if is_admin:
                    cv_headers = ['ID', 'User', 'Ngày đánh giá', 'Tên ứng viên', 'Email', 'Vị trí', 'Điểm', 'Đạt yêu cầu', 'Đã gửi email']
                else:
                    cv_headers = ['ID', 'Ngày đánh giá', 'Tên ứng viên', 'Email', 'Vị trí', 'Điểm', 'Đạt yêu cầu', 'Đã gửi email']
                for col, header in enumerate(cv_headers, 1):
                    cell = ws_cv.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Data
                for row_num, cv in enumerate(evaluations, 2):
                    col_offset = 1 if is_admin else 0
                    ws_cv.cell(row=row_num, column=1, value=cv.get('id'))
                    if is_admin:
                        ws_cv.cell(row=row_num, column=2, value=cv.get('user_name', cv.get('user_email', '')))
                    ws_cv.cell(row=row_num, column=2 + col_offset, value=str(cv.get('created_at', ''))[:19])
                    ws_cv.cell(row=row_num, column=3 + col_offset, value=cv.get('candidate_name', ''))
                    ws_cv.cell(row=row_num, column=4 + col_offset, value=cv.get('candidate_email', ''))
                    ws_cv.cell(row=row_num, column=5 + col_offset, value=cv.get('job_title', ''))
                    ws_cv.cell(row=row_num, column=6 + col_offset, value=cv.get('overall_score', 0))
                    ws_cv.cell(row=row_num, column=7 + col_offset, value='Có' if cv.get('is_qualified') else 'Không')
                    ws_cv.cell(row=row_num, column=8 + col_offset, value='Có' if cv.get('email_sent') else 'Không')
                
                # Auto-fit columns
                for col in ws_cv.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = min(len(str(cell.value)), 50)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_cv.column_dimensions[column].width = adjusted_width
            
            # Statistics sheet
            ws_stats = wb.create_sheet('Thống kê' + (' (Toàn hệ thống)' if is_admin else ''))
            email_stats = self.get_email_statistics(user_id, is_admin=is_admin)
            cv_stats = self.get_cv_statistics(user_id, is_admin=is_admin)
            
            # Admin note
            if is_admin:
                ws_stats.cell(row=1, column=1, value='⚠️ DỮ LIỆU TOÀN HỆ THỐNG (ADMIN VIEW)').font = Font(bold=True, size=12, color='FF0000')
                start_row = 3
            else:
                start_row = 1
            
            # Email stats
            ws_stats.cell(row=start_row, column=1, value='THỐNG KÊ EMAIL').font = Font(bold=True, size=14)
            ws_stats.cell(row=start_row+1, column=1, value='Tổng số email đã gửi:')
            ws_stats.cell(row=start_row+1, column=2, value=email_stats.get('total_sent', 0))
            ws_stats.cell(row=start_row+2, column=1, value='Số email đã nhận phản hồi:')
            ws_stats.cell(row=start_row+2, column=2, value=email_stats.get('responded', 0))
            ws_stats.cell(row=start_row+3, column=1, value='Số email đang chờ:')
            ws_stats.cell(row=start_row+3, column=2, value=email_stats.get('pending', 0))
            ws_stats.cell(row=start_row+4, column=1, value='Tỷ lệ phản hồi:')
            ws_stats.cell(row=start_row+4, column=2, value=f"{email_stats.get('response_rate', 0)}%")
            
            # CV stats
            ws_stats.cell(row=start_row+6, column=1, value='THỐNG KÊ CV').font = Font(bold=True, size=14)
            ws_stats.cell(row=start_row+7, column=1, value='Tổng số CV đã đánh giá:')
            ws_stats.cell(row=start_row+7, column=2, value=cv_stats.get('total', 0))
            ws_stats.cell(row=start_row+8, column=1, value='Số ứng viên đạt yêu cầu:')
            ws_stats.cell(row=start_row+8, column=2, value=cv_stats.get('qualified', 0))
            ws_stats.cell(row=start_row+9, column=1, value='Điểm trung bình:')
            ws_stats.cell(row=start_row+9, column=2, value=cv_stats.get('average_score', 0))
            
            # Auto-fit
            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 15
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except ImportError:
            logger.error("openpyxl not installed")
            raise Exception("Thư viện openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            raise e
    
    def get_chart_data(self, user_id: int, chart_type: str = 'overview', is_admin: bool = False) -> Dict:
        """Get data formatted for specific chart types (all data if admin)"""
        email_stats = self.get_email_statistics(user_id, is_admin=is_admin)
        cv_stats = self.get_cv_statistics(user_id, is_admin=is_admin)
        
        if chart_type == 'pie':
            # Sentiment pie chart
            return {
                'type': 'pie',
                'title': '🎯 Phân tích cảm xúc phản hồi',
                'labels': ['Tích cực 😊', 'Tiêu cực 😞', 'Trung tính 😐'],
                'data': [
                    email_stats.get('sentiments', {}).get('positive', 0),
                    email_stats.get('sentiments', {}).get('negative', 0),
                    email_stats.get('sentiments', {}).get('neutral', 0)
                ],
                'colors': ['#10b981', '#ef4444', '#f59e0b']
            }
        
        elif chart_type == 'bar':
            # Top recipients bar chart
            top_recipients = email_stats.get('top_recipients', [])
            
            # Fallback if no recipients
            if not top_recipients:
                return {
                    'type': 'bar',
                    'title': '📊 Top người nhận email nhiều nhất',
                    'labels': ['Chưa có dữ liệu'],
                    'datasets': [{
                        'label': 'Email đã gửi',
                        'data': [0],
                        'backgroundColor': 'rgba(99, 102, 241, 0.8)',
                        'borderColor': '#6366f1',
                        'borderWidth': 2,
                        'borderRadius': 8
                    }]
                }
            
            # Get proper labels - use name if available, otherwise use email
            labels = []
            for r in top_recipients:
                name = r.get('name', '').strip()
                email = r.get('email', '').strip()
                if name:
                    labels.append(name[:20])
                elif email:
                    # Shorten email if too long
                    if len(email) > 20:
                        labels.append(email[:17] + '...')
                    else:
                        labels.append(email)
                else:
                    labels.append('Unknown')
            
            return {
                'type': 'bar',
                'title': '📊 Top người nhận email nhiều nhất',
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Email đã gửi',
                        'data': [r.get('count', 0) for r in top_recipients],
                        'backgroundColor': 'rgba(99, 102, 241, 0.8)',
                        'borderColor': '#6366f1',
                        'borderWidth': 2,
                        'borderRadius': 8
                    },
                    {
                        'label': 'Đã phản hồi',
                        'data': [r.get('responded', 0) for r in top_recipients],
                        'backgroundColor': 'rgba(16, 185, 129, 0.8)',
                        'borderColor': '#10b981',
                        'borderWidth': 2,
                        'borderRadius': 8
                    }
                ]
            }
        
        elif chart_type == 'line':
            # Emails by date line chart
            emails_by_date = email_stats.get('emails_by_date', [])
            return {
                'type': 'line',
                'title': '📈 Xu hướng email theo thời gian',
                'labels': [d.get('date', '') for d in emails_by_date[-30:]],
                'datasets': [
                    {
                        'label': 'Email đã gửi',
                        'data': [d.get('sent', 0) for d in emails_by_date[-30:]],
                        'borderColor': '#6366f1',
                        'backgroundColor': 'rgba(99, 102, 241, 0.1)',
                        'fill': True,
                        'tension': 0.4,
                        'pointRadius': 4,
                        'pointHoverRadius': 6
                    },
                    {
                        'label': 'Phản hồi nhận được',
                        'data': [d.get('responded', 0) for d in emails_by_date[-30:]],
                        'borderColor': '#10b981',
                        'backgroundColor': 'rgba(16, 185, 129, 0.1)',
                        'fill': True,
                        'tension': 0.4,
                        'pointRadius': 4,
                        'pointHoverRadius': 6
                    }
                ]
            }
        
        elif chart_type == 'doughnut':
            # Email status doughnut
            return {
                'type': 'doughnut',
                'title': '📬 Trạng thái email',
                'labels': ['Đã phản hồi ✅', 'Chờ phản hồi ⏳'],
                'data': [email_stats.get('responded', 0), email_stats.get('pending', 0)],
                'colors': ['#10b981', '#f59e0b']
            }
        
        elif chart_type == 'cv':
            # CV score distribution
            score_dist = cv_stats.get('score_distribution', {})
            return {
                'type': 'bar',
                'title': '📋 Phân bố điểm đánh giá CV',
                'labels': list(score_dist.keys()) if score_dist else ['0-40%', '40-60%', '60-80%', '80-100%'],
                'datasets': [{
                    'label': 'Số lượng CV',
                    'data': list(score_dist.values()) if score_dist else [0, 0, 0, 0],
                    'backgroundColor': [
                        'rgba(239, 68, 68, 0.8)',   # Red - low
                        'rgba(245, 158, 11, 0.8)',  # Orange - medium
                        'rgba(59, 130, 246, 0.8)',  # Blue - good
                        'rgba(16, 185, 129, 0.8)'   # Green - excellent
                    ],
                    'borderColor': ['#ef4444', '#f59e0b', '#3b82f6', '#10b981'],
                    'borderWidth': 2,
                    'borderRadius': 8
                }]
            }
        
        elif chart_type == 'decision':
            # Decision analysis
            decisions = email_stats.get('decisions', {})
            return {
                'type': 'polarArea',
                'title': '🎯 Phân tích quyết định từ phản hồi',
                'labels': ['Đồng ý ✅', 'Từ chối ❌', 'Chưa quyết định 🤔', 'Cần thêm thông tin ℹ️'],
                'data': [
                    decisions.get('agreed', 0),
                    decisions.get('disagreed', 0),
                    decisions.get('undecided', 0),
                    decisions.get('needs_more_info', 0)
                ],
                'colors': ['#10b981', '#ef4444', '#f59e0b', '#3b82f6']
            }
        
        elif chart_type == 'cv_qualified':
            # CV qualification pie
            return {
                'type': 'pie',
                'title': '👥 Tỷ lệ ứng viên đạt yêu cầu',
                'labels': ['Đạt yêu cầu (≥85%) ✅', 'Chưa đạt ❌'],
                'data': [cv_stats.get('qualified', 0), cv_stats.get('not_qualified', 0)],
                'colors': ['#10b981', '#ef4444']
            }
        
        elif chart_type == 'radar':
            # Overall performance radar
            total_emails = email_stats.get('total_sent', 0) or 1
            total_cv = cv_stats.get('total', 0) or 1
            
            return {
                'type': 'radar',
                'title': '📊 Tổng quan hiệu suất',
                'labels': ['Tỷ lệ phản hồi', 'Phản hồi tích cực', 'Tỷ lệ CV đạt', 'Email thành công', 'Đánh giá CV'],
                'datasets': [{
                    'label': 'Hiệu suất (%)',
                    'data': [
                        email_stats.get('response_rate', 0),
                        (email_stats.get('sentiments', {}).get('positive', 0) / max(email_stats.get('responded', 1), 1)) * 100,
                        cv_stats.get('qualification_rate', 0),
                        min((email_stats.get('responded', 0) / total_emails) * 100, 100),
                        min((cv_stats.get('qualified', 0) / total_cv) * 100, 100)
                    ],
                    'backgroundColor': 'rgba(99, 102, 241, 0.2)',
                    'borderColor': '#6366f1',
                    'pointBackgroundColor': '#6366f1',
                    'pointBorderColor': '#fff',
                    'pointHoverBackgroundColor': '#fff',
                    'pointHoverBorderColor': '#6366f1'
                }]
            }
        
        else:  # overview - email status doughnut
            return {
                'type': 'doughnut',
                'title': '📬 Tổng quan trạng thái email',
                'labels': ['Đã phản hồi ✅', 'Chờ phản hồi ⏳'],
                'data': [email_stats.get('responded', 0), email_stats.get('pending', 0)],
                'colors': ['#10b981', '#f59e0b']
            }
